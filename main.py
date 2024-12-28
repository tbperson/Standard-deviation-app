import tkinter as tk
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import filedialog

class TableApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Standard_Deviation")
        self.configure(bg="green")  # Set background color for the entire application
        self.create_widgets()

    def create_widgets(self):
        tk.Label(self, text="Input n (number of values):").pack()
        self.rows_entry = tk.Entry(self)
        self.rows_entry.pack()

        self.create_button = tk.Button(self, text="Create Table", command=self.create_table)
        self.create_button.pack()

        self.calculate_button = tk.Button(self, text="Calculate the standard deviation", command=self.calculate_standard_deviation, state=tk.DISABLED)
        self.calculate_button.pack()

        self.export_button = tk.Button(self, text="Export data", command=self.export_data)
        self.export_button.pack()

        # Add a text widget to display the standard deviation
        self.rank_display = tk.Text(self, height=2, width=30)
        self.rank_display.pack()

    def create_table(self):
        self.create_button.config(state=tk.DISABLED)
        num_input_columns = 1

        try:
            self.num_rows = int(self.rows_entry.get())
        except ValueError:
            quit()
            return

        # Define table_canvas within the create_table method
        table_canvas = tk.Canvas(self, bg="green")  # Set background color for the table canvas
        table_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        table_frame = tk.Frame(table_canvas, bg="green")  # Set background color for the table frame
        table_canvas.create_window((0, 0), window=table_frame, anchor=tk.NW)

        self.input_entry_vars = []
        for i in range(self.num_rows):
            entry_row = []
            for j in range(num_input_columns):
                entry_var = tk.DoubleVar()
                entry = tk.Entry(table_frame, textvariable=entry_var, borderwidth=1, relief="solid", width=15)
                entry.grid(row=i, column=j, padx=1, pady=1, sticky="w")
                entry_row.append(entry_var)
            self.input_entry_vars.append(entry_row)

        self.input_entry_vars2 = []
        for i in range(self.num_rows):
            entry_row2 = []
            for j in range(1):
                entry_var2 = tk.DoubleVar()
                entry2 = tk.Entry(table_frame, textvariable=entry_var2, borderwidth=1, relief="solid", width=15)
                entry2.grid(row=i, column=j+num_input_columns, padx=1, pady=1, sticky="w")
                entry_row2.append(entry_var2)
            self.input_entry_vars2.append(entry_row2)

        table_frame.update_idletasks()
        table_canvas.config(scrollregion=table_canvas.bbox("all"))

        v_scrollbar = tk.Scrollbar(self, orient=tk.VERTICAL, command=table_canvas.yview)
        v_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        table_canvas.configure(yscrollcommand=v_scrollbar.set)

        self.calculate_button.config(state=tk.NORMAL)
        self.rank_display.insert(tk.END, "Insert data into column 1") 

    def get_mean(self):
        data = []
        for i in range(self.num_rows):
            entry_var = self.input_entry_vars[i][0]
            value = entry_var.get()
            data.append(value)
        mean = sum(data) / len(data)
        return mean

    def get_differences(self, mean):
        differences_list = []
        for i in range(self.num_rows):
            entry_var = self.input_entry_vars[i][0]
            value = float(entry_var.get())
            difference = value - mean
            differences_list.append(difference)
            
        return differences_list
    
    def update_differences(self, differences_list):
        for i in range(self.num_rows):
            entry_var = self.input_entry_vars2[i][0]
            entry_var.set(differences_list[i])

    def calculate_standard_deviation(self):
        mean = self.get_mean()
        differences = self.get_differences(mean)
        self.update_differences(differences)
        variance = sum([diff ** 2 for diff in differences]) / len(differences)
        standard_deviation = round(variance ** 0.5, 4)
        self.rank_display.delete(1.0, tk.END)
        self.rank_display.insert(tk.END, f"Mean: {mean}\n")
        self.rank_display.insert(tk.END, f"Standard Deviation: {standard_deviation}")

    def export_data(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "Standard deviation data"
        values = [self.input_entry_vars[i][0].get() for i in range(self.num_rows)]
        differences = [self.input_entry_vars2[i][0].get() for i in range(self.num_rows)]
        mean = self.get_mean()
        differences_from_mean = self.get_differences(mean)
        variance = sum([diff ** 2 for diff in differences_from_mean]) / len(differences_from_mean)
        standard_deviation = round(variance ** 0.5, 4)
        
        header = (["Index", "Value 1", "Value 2", "D1", "D2"])
        ws.append(header)
        

        for i, (value1, value2, diff1, diff2) in enumerate(zip(values, differences, differences_from_mean, differences)):
            ws.append([i + 1, value1, value2, diff1, diff2])
        
        ws.append([])
        ws.append(["Mean", mean])
        ws.append(["Standard Deviation", standard_deviation])

        green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
        black_border = Border(left=Side(style='thin', color='000000'),
                          right=Side(style='thin', color='000000'),
                          top=Side(style='thin', color='000000'),
                          bottom=Side(style='thin', color='000000'))

        for cell in ws[1]:
            cell.fill = green_fill

        for row in ws.iter_rows():
            for cell in row:
                cell.border = black_border

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")


        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if file_path:
            wb.save(file_path)

if __name__ == "__main__":
    app = TableApp()
    app.mainloop()
